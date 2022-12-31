from tkinter import *
from tkinter import filedialog , ttk
from openpyxl import Workbook , styles

master = Tk()
master.geometry("700x700")
master.configure(bg='#000000')

# To capture the mouse coordinates
# Just to arrange the frames in the suitable place
def coordinates(event):
	print("Mouse Coordinates: " + str(event.x) + " , " + str(event.y))

master.bind("<Button-1>", coordinates) # Left mouse click

var_list = ["Averages" , "RTWP" , "Benchmark"]

frame_select_file = Frame(master , bd = 10 , bg = "#CA6B45")
frame_select_file.place(x=10, y=100)

label_select_file = Label(frame_select_file  , fg = "blue" , text="Select file: " , bd = 10)
label_select_file.pack(side=LEFT)

entry_select_file = Entry(frame_select_file, width = 70 , bd = 10)
entry_select_file.pack(side=RIGHT)

center_Alignment = styles.Alignment(horizontal="center", vertical="center", wrapText=True)

###############################################################################################################################
###############################################################################################################################

def averegas_code():
    files_path = filedialog.askopenfilenames(initialdir="/", 
                                             title="Select files" , 
                                             filetypes=(("FMT files", "*.FMT"),("all files", "*.*"))
                                            )

    entry_select_file.delete(0,200)
    entry_select_file.insert(0,str(files_path))
    print(files_path)
    print(type(files_path))

    for path in files_path :
        print("")
        print("=================================================================================")
        print(f"==============================={path}=============================================")
        print("=================================================================================")
        print("")

        wb = Workbook()
        sheet = wb.active 

        window = Toplevel()
        window.geometry('600x700')

        newlabel = Label(window, text = path)
        newlabel.pack()

        window.title(path)
        output = Text(window , font = "Times 14 bold" , width= 500, height= 300)

        file = open(path).readlines()
        print("Opened!")

        col_to_extract = [7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22]

        print(col_to_extract)
        print(len(file))

        ci_list = []
        as_sc_list = []
        san_sc_list = []
        pci_list = []

        ci_count = [0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 ]
        as_sc_count = [0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 ]
        san_sc_count = [0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 ]
        pci_count = [0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 ]

        for i in col_to_extract:
            avg = sum1 = count = max1 = 0
            for r in range(1 , len(file)):
                value = file[r].split("\t")[i]
                if value != '':
                    try :
                        value_float = float(value)
                    except:
                        continue

                    sum1 = value_float + sum1
                    count = count + 1
                    if i == 11 or i == 12 or i == 15 or i == 16 :
                        if value_float >= max1:
                            max1 = value_float 

                    elif i == 17 :
                        if value_float not in ci_list :
                            ci_list.append(value_float)

                    elif i == 18 :
                        if value_float not in as_sc_list :
                            as_sc_list.append(value_float)

                    elif i == 22 :
                        if value_float not in san_sc_list :
                            san_sc_list.append(value_float)

                    elif i == 19 :
                        if value_float not in pci_list :
                            pci_list.append(value_float)

            if count != 0:
                avg = sum1/count

            if i == 11 or i == 12 or i == 15 or i == 16 :
                print("average " + file[0].split("\t")[i] + " = " + str(round(avg,2)))
                print("max " + file[0].split("\t")[i] + " = " + str(round(max1,2)))
                output.insert(INSERT, "average " + file[0].split("\t")[i] + " = " + str(round(avg,2)) + "\n")
                output.insert(INSERT, "max " + file[0].split("\t")[i] + " = " + str(round(max1,2)) + "\n")

                c = sheet.cell(row= 1 , column = i-6) 
                c.value = "average " + file[0].split("\t")[i]

                c = sheet.cell(row= 2 , column = i-6) 
                c.value = round(avg,2)

                c = sheet.cell(row= 3 , column = i-6) 
                c.value = "max " + file[0].split("\t")[i] 

                c = sheet.cell(row= 4 , column = i-6) 
                c.value = round(max1,2)

            elif i == 17 :
                print(file[0].split("\t")[i] + " = " + str(ci_list))
                output.insert(INSERT, file[0].split("\t")[i] + " = " + str(ci_list) + "\n")

            elif i == 18 :
                print(file[0].split("\t")[i] + " = " + str(as_sc_list))
                output.insert(INSERT, file[0].split("\t")[i] + " = " + str(as_sc_list) + "\n") 

            elif i == 22 :
                print(file[0].split("\t")[i] + " = " + str(san_sc_list))
                output.insert(INSERT, file[0].split("\t")[i] + " = " + str(san_sc_list) + "\n")

            elif i == 19 :
                print(file[0].split("\t")[i] + " = " + str(pci_list))
                output.insert(INSERT, file[0].split("\t")[i] + " = " + str(pci_list) + "\n")

            else :
                print("average " + file[0].split("\t")[i] + " = " + str(round(avg,2)))
                output.insert(INSERT, "average " + file[0].split("\t")[i] + " = " + str(round(avg,2)) + "\n")

                c = sheet.cell(row= 1 , column = i-6) 
                c.value = "average " + file[0].split("\t")[i] 

                c = sheet.cell(row= 2 , column = i-6) 
                c.value = round(avg,2)

            if i == 8 or i == 12 or i == 16 or i == 19 or i == 22 :
                print()
                output.insert(INSERT,"\n")

        ###################################################################################################
        ####################################################################################################
        #######################################################################################################

        i = 0
        for i in [17,18,19,22]:
            for r in range(1 , len(file)):
                value = file[r].split("\t")[i]
                if value != '':
                    value_float = float(value)

                    if i == 17 :
                        for n in range(0 , len(ci_list)) :
                            if ci_list[n] == value_float :
                                ci_count[n] = ci_count[n] + 1
                    
                    elif i == 18 :
                        for n in range(0 , len(as_sc_list)) :
                            if as_sc_list[n] == value_float :
                                as_sc_count[n] = as_sc_count[n] + 1

                    elif i == 22 :
                        for n in range(0 , len(san_sc_list)) :
                            if san_sc_list[n] == value_float :
                                san_sc_count[n] = san_sc_count[n] + 1

                    elif i == 19 :
                        for n in range(0 , len(pci_list)) :
                            if pci_list[n] == value_float :
                                pci_count[n] = pci_count[n] + 1

        for n in range(0 , len(ci_list)) :
            print("CI " + str(ci_list[n]) + " Count = " + str(ci_count[n]))
            output.insert(INSERT, "CI " + str(ci_list[n]) + " Count = " + str(ci_count[n]) + "\n")

            c = sheet.cell(row= 5 , column = n+1) 
            c.value = "CI " + str(ci_list[n]) + " Count "

            c = sheet.cell(row= 6 , column = n+1) 
            c.value = ci_count[n]

        for n in range(0 , len(as_sc_list)) :
            print("AS SC " + str(as_sc_list[n]) + " Count = " + str(as_sc_count[n]) + "\n")
            output.insert(INSERT, "AS SC " + str(as_sc_list[n]) + " Count = " + str(as_sc_count[n]) + "\n")

            c = sheet.cell(row= 8 , column = n+1) 
            c.value = "AS SC " + str(as_sc_list[n]) + " Count "

            c = sheet.cell(row= 9 , column = n+1) 
            c.value = as_sc_count[n]
        
        for n in range(0 , len(san_sc_list)) :
            print("SAN SC " + str(san_sc_list[n]) + " Count = " + str(san_sc_count[n]) + "\n")
            output.insert(INSERT, "SAN SC " + str(san_sc_list[n]) + " Count = " + str(san_sc_count[n]) + "\n")

            c = sheet.cell(row= 11 , column = n+1) 
            c.value = "SAN SC " + str(san_sc_list[n]) + " Count "

            c = sheet.cell(row= 12 , column = n+1) 
            c.value = san_sc_count[n]
        
        for n in range(0 , len(pci_list)) :
            print("PCI " + str(pci_list[n]) + " Count = " + str(pci_count[n]))
            output.insert(INSERT, "PCI " + str(pci_list[n]) + " Count = " + str(pci_count[n]) + "\n")

            c = sheet.cell(row= 14 , column = n+1) 
            c.value = "PCI " + str(pci_list[n]) + " Count "

            c = sheet.cell(row= 15 , column = n+1) 
            c.value = pci_count[n]

        print("")
        print("=================================================================================")
        print("=======================================")
        print("=================================================================================")
        print("")

        output.pack()
        for row in range(1,33):
            for col in range(1,33):
                sheet.cell(row,col).alignment = center_Alignment
        wb.save(filename=f"{path}.xlsx")
    
def rtwp_code():
    files_path = filedialog.askopenfilenames(initialdir="/", 
                                        title="Select files" , 
                                        filetypes=(("FMT files", "*.FMT"),("all files", "*.*"))
                                       )
    entry_select_file.delete(0,200)
    entry_select_file.insert(0,str(files_path))
    print(files_path)
    print(type(files_path))

    for path in files_path :
        print("")
        print("=================================================================================")
        print(f"==============================={path}=============================================")
        print("=================================================================================")
        print("")

        wb = Workbook()
        sheet = wb.active 

        window = Toplevel()
        window.geometry('600x700')

        newlabel = Label(window, text = path)
        newlabel.pack()

        window.title(path)
        output = Text(window , font = "Times 14 bold" , width= 500, height= 300)

        file = open(path).readlines()
        print("Opened!")
        col_to_extract = [7,8,9,10,11,12,13]

        print(col_to_extract)
        print(len(file))

        as_sc_list = []
        san_sc_list = []

        interference_int_value = -200
        interference_list = dict()

        as_sc_count = [0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 ]
        san_sc_count = [0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 ]

        for i in col_to_extract:
            avg = sum1 = count = 0
            for r in range(1 , len(file)):
                value = file[r].split("\t")[i]
                if value != '':
                    try :
                        value_float = float(value)
                    except:
                        continue

                    sum1 = value_float + sum1
                    count = count + 1

                    if i == 9 :
                        if value_float not in as_sc_list :
                            as_sc_list.append(value_float)


                    if i == 12 :
                        if value_float not in san_sc_list :
                            san_sc_list.append(value_float)

            if count != 0:
                avg = sum1/count

            if i == 9 :
                print(file[0].split("\t")[i] + " = " + str(as_sc_list))
                output.insert(INSERT, file[0].split("\t")[i] + " = " + str(as_sc_list) + "\n")

            elif i == 12 :
                print(file[0].split("\t")[i] + " = " + str(san_sc_list))
                output.insert(INSERT, file[0].split("\t")[i] + " = " + str(san_sc_list) + "\n")

            else :
                print("Average " + file[0].split("\t")[i] + " = " + str(round(avg,2)))
                output.insert(INSERT, "Average " + file[0].split("\t")[i] + " = " + str(round(avg,2)) + "\n")

                c = sheet.cell(row= 1 , column = i-6) 
                c.value = "Average " + file[0].split("\t")[i]

                c = sheet.cell(row= 2 , column = i-6) 
                c.value = round(avg,2)

            if i == 9 or i == 12 :
                print()
                output.insert(INSERT,"\n")
            
        ###################################################################################################
        ####################################################################################################
        #######################################################################################################

        for v in san_sc_list:
            interference_list[f'{v}'] = interference_int_value
        print(interference_list)

        i = 0
        for i in [9,12]:
            for r in range(1 , len(file)):
                value = file[r].split("\t")[i]
                if value != '':
                    value_float = float(value)

                    if i == 9 :
                        for n in range(0 , len(as_sc_list)) :
                            if as_sc_list[n] == value_float :
                                as_sc_count[n] = as_sc_count[n] + 1

                    elif i == 12 :
                        for n in range(0 , len(san_sc_list)) :
                            if san_sc_list[n] == value_float :
                                san_sc_count[n] = san_sc_count[n] + 1

                        interference = file[r].split("\t")[13]
                        if interference != '':
                            interference = float(file[r].split("\t")[13])
                            
                            if  interference > float(interference_list[f'{value_float}']):
                                interference_list[f'{value_float}'] = interference  

        for n in range(0 , len(as_sc_list)) :
            print("AS SC " + str(as_sc_list[n]) + " Count = " + str(as_sc_count[n]))
            output.insert(INSERT, "AS SC " + str(as_sc_list[n]) + " Count = " + str(as_sc_count[n]) + "\n")

            c = sheet.cell(row= 5 , column = n+1) 
            c.value = "AS SC " + str(as_sc_list[n]) + " Count "

            c = sheet.cell(row= 6 , column = n+1) 
            c.value = as_sc_count[n]

        for n in range(0 , len(san_sc_list)) :
            print("SAN SC " + str(san_sc_list[n]) + " Count = " + str(san_sc_count[n]))
            output.insert(INSERT, "SAN SC " + str(san_sc_list[n]) + " Count = " + str(san_sc_count[n]) + "\n")

            c = sheet.cell(row= 8 , column = n+1) 
            c.value = "SAN SC " + str(san_sc_list[n]) + " Count "

            c = sheet.cell(row= 9 , column = n+1) 
            c.value = san_sc_count[n]

        output.insert(INSERT, "Interference List = " + str(interference_list) + "\n")
        print(interference_list) 

        p = 0
        for key , value1 in interference_list.items():
            output.insert(INSERT, "SC " + str(key) + " interference = " + str(value1) + "\n") 
            
            c = sheet.cell(row= 1 , column = 8 + p) 
            c.value = "SC " + str(key) + " interference "

            c = sheet.cell(row= 2 , column = 8 + p) 
            c.value = value1
            p = p + 1

        print("")
        print("=================================================================================")
        print("=======================================")
        print("=================================================================================")
        print("")

        output.pack()
        for row in range(1,33):
            for col in range(1,33):
                sheet.cell(row,col).alignment = center_Alignment
        wb.save(filename=f"{path}.xlsx")

###############################################################################################################################
###############################################################################################################################
###############################################################################################################################
###############################################################################################################################
###############################################################################################################################

def benchmark_code():
    files_path = filedialog.askopenfilenames(initialdir="/", 
                                        title="Select files" , 
                                        filetypes=(("FMT files", "*.FMT"),("all files", "*.*"))
                                       )
    entry_select_file.delete(0,200)
    entry_select_file.insert(0,str(files_path))
    print(files_path)
    print(type(files_path))


    for path in files_path :
        print("")
        print("=================================================================================")
        print(f"==============================={path}=============================================")
        print("=================================================================================")
        print("")

        wb = Workbook()
        sheet = wb.active 

        window = Toplevel()
        window.geometry('600x700')

        newlabel = Label(window, text = path)
        newlabel.pack()

        window.title(path)
        output = Text(window , font = "Times 14 bold" , width= 500, height= 300)

        file = open(path).readlines()
        print("Opened!") # To make sure

        col_to_extract = [7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22]
        col_of_eq = 1

        print(col_to_extract)
        print(len(file))

        ci_list_eq1 = []
        as_sc_list_eq1 = []
        san_sc_list_eq1 = []
        pci_list_eq1 = []

        ci_count_eq1 = [0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 ]
        as_sc_count_eq1 = [0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 ]
        san_sc_count_eq1 = [0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 ]
        pci_count_eq1 = [0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 ]

        ci_list_eq2 = []
        as_sc_list_eq2 = []
        san_sc_list_eq2 = []
        pci_list_eq2 = []

        ci_count_eq2 = [0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 ]
        as_sc_count_eq2 = [0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 ]
        san_sc_count_eq2 = [0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 ]
        pci_count_eq2 = [0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 , 0 ]

        
        for i in col_to_extract:
            avg = avg2 = sum1 = sum2 = count = count2 = max1 = max2 = 0
            for r in range(1 , len(file)):
                eq = file[r].split("\t")[col_of_eq]
                
                if eq == "EQ1" :
                    value1 = file[r].split("\t")[i]
                    if value1 != '':
                        try :
                            value_float1 = float(value1)
                        except:
                            continue

                        sum1 = value_float1 + sum1
                        count = count + 1
                        if i == 11 or i == 12 or i == 15 or i == 16 :
                            if value_float1 >= max1:
                                max1 = value_float1

                        elif i == 17 :
                            if value_float1 not in ci_list_eq1 :
                                ci_list_eq1.append(value_float1)

                        elif i == 18 :
                            if value_float1 not in as_sc_list_eq1 :
                                as_sc_list_eq1.append(value_float1)

                        elif i == 22 :
                            if value_float1 not in san_sc_list_eq1 :
                                san_sc_list_eq1.append(value_float1)

                        elif i == 19 :
                            if value_float1 not in pci_list_eq1 :
                                pci_list_eq1.append(value_float1)

                elif eq == "EQ2" :
                    value2 = file[r].split("\t")[i]
                    if value2 != '':
                        try :
                            value_float2 = float(value2)
                        except:
                            continue

                        sum2 = value_float2 + sum2
                        count2 = count2 + 1
                        if i == 11 or i == 12 or i == 15 or i == 16 :
                            if value_float2 >= max2:
                                max2 = value_float2 

                        elif i == 17 :
                            if value_float2 not in ci_list_eq2 :
                                ci_list_eq2.append(value_float2)

                        elif i == 18 :
                            if value_float2 not in as_sc_list_eq2 :
                                as_sc_list_eq2.append(value_float2)

                        elif i == 22 :
                            if value_float2 not in san_sc_list_eq2 :
                                san_sc_list_eq2.append(value_float2)

                        elif i == 19 :
                            if value_float2 not in pci_list_eq2 :
                                pci_list_eq2.append(value_float2)

            if count != 0 :
                avg = sum1/count
                
            if count2 != 0 :
                avg2 = sum2/count2

            if i == 11 or i == 12 or i == 15 or i == 16 :
                print("average EQ1 " + file[0].split("\t")[i] + " = " + str(round(avg,2)))
                print("max  EQ1 " + file[0].split("\t")[i] + " = " + str(round(max1,2)))
                output.insert(INSERT, "average  EQ1 " + file[0].split("\t")[i] + " = " + str(round(avg,2)) + "\n")
                output.insert(INSERT, "max  EQ1 " + file[0].split("\t")[i] + " = " + str(round(max1,2)) + "\n")

                c = sheet.cell(row= 1 , column = i-6) 
                c.value = "average  EQ1 " + file[0].split("\t")[i]

                c = sheet.cell(row= 2 , column = i-6) 
                c.value = round(avg,2)

                c = sheet.cell(row= 3 , column = i-6) 
                c.value = "max  EQ1 " + file[0].split("\t")[i]
                
                c = sheet.cell(row= 4 , column = i-6) 
                c.value = round(max1,2)

                print("average  EQ2 " + file[0].split("\t")[i] + " = " + str(round(avg2,2)))
                print("max  EQ2 " + file[0].split("\t")[i] + " = " + str(round(max2,2)))
                output.insert(INSERT, "average  EQ2 " + file[0].split("\t")[i] + " = " + str(round(avg2,2)) + "\n")
                output.insert(INSERT, "max  EQ2 " + file[0].split("\t")[i] + " = " + str(round(max2,2)) + "\n")

                c = sheet.cell(row= 6 , column = i-6) 
                c.value = "average  EQ2 " + file[0].split("\t")[i]

                c = sheet.cell(row= 7 , column = i-6) 
                c.value = round(avg2,2)

                c = sheet.cell(row= 8 , column = i-6) 
                c.value = "max  EQ2 " + file[0].split("\t")[i]
                
                c = sheet.cell(row= 9 , column = i-6) 
                c.value = round(max2,2)

                


            elif i == 17 :
                print(file[0].split("\t")[i] + " = " + str(ci_list_eq1))
                output.insert(INSERT, "EQ1 " + file[0].split("\t")[i] + " = " + str(ci_list_eq1) + "\n")

                print(file[0].split("\t")[i] + " = " + str(ci_list_eq2))
                output.insert(INSERT, "EQ2 " + file[0].split("\t")[i] + " = " + str(ci_list_eq2) + "\n")


            elif i == 18 :
                print(file[0].split("\t")[i] + " = " + str(as_sc_list_eq1))
                output.insert(INSERT, "EQ1 " + file[0].split("\t")[i] + " = " + str(as_sc_list_eq1) + "\n") 

                print(file[0].split("\t")[i] + " = " + str(as_sc_list_eq2))
                output.insert(INSERT, "EQ2 " + file[0].split("\t")[i] + " = " + str(as_sc_list_eq2) + "\n") 

            elif i == 22 :
                print(file[0].split("\t")[i] + " = " + str(san_sc_list_eq1))
                output.insert(INSERT, "EQ1 " + file[0].split("\t")[i] + " = " + str(san_sc_list_eq1) + "\n")

                print(file[0].split("\t")[i] + " = " + str(san_sc_list_eq2))
                output.insert(INSERT, "EQ2 " + file[0].split("\t")[i] + " = " + str(san_sc_list_eq2) + "\n")

            elif i == 19 :
                print(file[0].split("\t")[i] + " = " + str(pci_list_eq1))
                output.insert(INSERT, "EQ1 " + file[0].split("\t")[i] + " = " + str(pci_list_eq1) + "\n")

                print(file[0].split("\t")[i] + " = " + str(pci_list_eq2))
                output.insert(INSERT, "EQ2 " + file[0].split("\t")[i] + " = " + str(pci_list_eq2) + "\n")

            else :
                print("average " + file[0].split("\t")[i] + " = " + str(round(avg,2)))
                output.insert(INSERT, "EQ1 average " + file[0].split("\t")[i] + " = " + str(round(avg,2)) + "\n")

                c = sheet.cell(row= 1 , column = i-6) 
                c.value = "average  EQ1 " + file[0].split("\t")[i]

                c = sheet.cell(row= 2 , column = i-6) 
                c.value = round(avg,2)

                print("average " + file[0].split("\t")[i] + " = " + str(round(avg2,2)))
                output.insert(INSERT, "EQ2 average " + file[0].split("\t")[i] + " = " + str(round(avg2,2)) + "\n")

                c = sheet.cell(row= 4 , column = i-6) 
                c.value = "average  EQ2 " + file[0].split("\t")[i]

                c = sheet.cell(row= 5 , column = i-6) 
                c.value = round(avg2,2)

            if i == 8 or i == 12 or i == 16 or i == 19 or i == 22 :
                print()
                output.insert(INSERT,"\n")

        ###################################################################################################
        ####################################################################################################
        #######################################################################################################

        i = 0
        for i in [17,18,19,22]:
            for r in range(1 , len(file)):
                eq = file[r].split("\t")[col_of_eq]
                value = file[r].split("\t")[i]

                if eq == "EQ1" :
                    if value != '':
                        value_float = float(value)

                        if i == 17 :
                            for n in range(0 , len(ci_list_eq1)) :
                                if ci_list_eq1[n] == value_float :
                                    ci_count_eq1[n] = ci_count_eq1[n] + 1

                        elif i == 18 :
                            for n in range(0 , len(as_sc_list_eq1)) :
                                if as_sc_list_eq1[n] == value_float :
                                    as_sc_count_eq1[n] = as_sc_count_eq1[n] + 1

                        elif i == 22 :
                            for n in range(0 , len(san_sc_list_eq1)) :
                                if san_sc_list_eq1[n] == value_float :
                                    san_sc_count_eq1[n] = san_sc_count_eq1[n] + 1

                        elif i == 19 :
                            for n in range(0 , len(pci_list_eq1)) :
                                if pci_list_eq1[n] == value_float :
                                    pci_count_eq1[n] = pci_count_eq1[n] + 1
    
                elif eq == "EQ2" :
                    if value != '':
                        value_float = float(value)

                        if i == 17 :
                            for n in range(0 , len(ci_list_eq2)) :
                                if ci_list_eq2[n] == value_float :
                                    ci_count_eq2[n] = ci_count_eq2[n] + 1

                        elif i == 18 :
                            for n in range(0 , len(as_sc_list_eq2)) :
                                if as_sc_list_eq2[n] == value_float :
                                    as_sc_count_eq2[n] = as_sc_count_eq2[n] + 1

                        elif i == 22 :
                            for n in range(0 , len(san_sc_list_eq2)) :
                                if san_sc_list_eq2[n] == value_float :
                                    san_sc_count_eq2[n] = san_sc_count_eq2[n] + 1

                        elif i == 19 :
                            for n in range(0 , len(pci_list_eq2)) :
                                if pci_list_eq2[n] == value_float :
                                    pci_count_eq2[n] = pci_count_eq2[n] + 1

        for n in range(0 , len(ci_list_eq1)) :
            print("CI " + str(ci_list_eq1[n]) + " Count = " + str(ci_count_eq1[n]))
            output.insert(INSERT, "EQ1 CI " + str(ci_list_eq1[n]) + " Count = " + str(ci_count_eq1[n]) + "\n")

            c = sheet.cell(row= 11 , column = n+1) 
            c.value = "EQ1 CI " + str(ci_list_eq1[n]) + " Count "

            c = sheet.cell(row= 12 , column = n+1) 
            c.value = ci_count_eq1[n]

        for n in range(0 , len(as_sc_list_eq1)) :
            print("SC " + str(as_sc_list_eq1[n]) + " Count = " + str(as_sc_count_eq1[n]))
            output.insert(INSERT, "EQ1 AS SC " + str(as_sc_list_eq1[n]) + " Count = " + str(as_sc_count_eq1[n]) + "\n")

            c = sheet.cell(row= 14 , column = n+1) 
            c.value = "EQ1 AS SC " + str(as_sc_list_eq1[n]) + " Count "

            c = sheet.cell(row= 15 , column = n+1) 
            c.value = as_sc_count_eq1[n]
        
        for n in range(0 , len(san_sc_list_eq1)) :
            print("SC " + str(san_sc_list_eq1[n]) + " Count = " + str(san_sc_count_eq1[n]))
            output.insert(INSERT, "EQ1 SAN SC " + str(san_sc_list_eq1[n]) + " Count = " + str(san_sc_count_eq1[n]) + "\n")

            c = sheet.cell(row= 17 , column = n+1) 
            c.value = "EQ1 SAN SC " + str(san_sc_list_eq1[n]) + " Count "

            c = sheet.cell(row= 18 , column = n+1) 
            c.value = san_sc_count_eq1[n]
        
        for n in range(0 , len(pci_list_eq1)) :
            print("PCI " + str(pci_list_eq1[n]) + " Count = " + str(pci_count_eq1[n]))
            output.insert(INSERT, "EQ1 PCI " + str(pci_list_eq1[n]) + " Count = " + str(pci_count_eq1[n]) + "\n")

            c = sheet.cell(row= 21 , column = n+1) 
            c.value = "EQ1 PCI " + str(pci_list_eq1[n]) + " Count "

            c = sheet.cell(row= 22 , column = n+1) 
            c.value = pci_count_eq1[n]

################################################################################################################################

        for n in range(0 , len(ci_list_eq2)) :
            print("CI " + str(ci_list_eq2[n]) + " Count = " + str(ci_count_eq2[n]))
            output.insert(INSERT, "EQ2 CI " + str(ci_list_eq2[n]) + " Count = " + str(ci_count_eq2[n]) + "\n")

            c = sheet.cell(row= 23 , column = n+1) 
            c.value = "EQ2 CI " + str(ci_list_eq2[n]) + " Count "

            c = sheet.cell(row= 24 , column = n+1) 
            c.value = ci_count_eq2[n]

        for n in range(0 , len(as_sc_list_eq2)) :
            print("SC " + str(as_sc_list_eq2[n]) + " Count = " + str(as_sc_count_eq2[n]))
            output.insert(INSERT, "EQ2 AS SC " + str(as_sc_list_eq2[n]) + " Count = " + str(as_sc_count_eq2[n]) + "\n")

            c = sheet.cell(row= 26 , column = n+1) 
            c.value = "EQ2 AS SC " + str(as_sc_list_eq2[n]) + " Count "

            c = sheet.cell(row= 27 , column = n+1) 
            c.value = as_sc_count_eq2[n]
        
        for n in range(0 , len(san_sc_list_eq2)) :
            print("SC " + str(san_sc_list_eq2[n]) + " Count = " + str(san_sc_count_eq2[n]))
            output.insert(INSERT, "EQ2 SAN SC " + str(san_sc_list_eq2[n]) + " Count = " + str(san_sc_count_eq2[n]) + "\n")

            c = sheet.cell(row= 29 , column = n+1) 
            c.value = "EQ2 SAN SC " + str(san_sc_list_eq2[n]) + " Count "

            c = sheet.cell(row= 30 , column = n+1) 
            c.value = san_sc_count_eq2[n]
        
        for n in range(0 , len(pci_list_eq2)) :
            print("PCI " + str(pci_list_eq2[n]) + " Count = " + str(pci_count_eq2[n]))
            output.insert(INSERT, "EQ2 PCI " + str(pci_list_eq2[n]) + " Count = " + str(pci_count_eq2[n]) + "\n")

            c = sheet.cell(row= 32 , column = n+1) 
            c.value = "EQ2 PCI " + str(pci_list_eq2[n]) + " Count "

            c = sheet.cell(row= 33 , column = n+1) 
            c.value = pci_count_eq2[n]

        print("")
        print("=================================================================================")
        print("=======================================")
        print("=================================================================================")
        print("")

        output.pack()
        for row in range(1,33):
            for col in range(1,33):
                sheet.cell(row,col).alignment = center_Alignment
        wb.save(filename=f"{path}.xlsx")


###############################################################################################################################
###############################################################################################################################
###############################################################################################################################
###############################################################################################################################
###############################################################################################################################


def retrieve() :
    option = combo.get()
    print("Your Choice is : " + option)
    print("\n\n\n")
    if option == "Averages" :
        averegas_code()
    elif option == "RTWP" :
        rtwp_code()
    elif option == "Benchmark":
        benchmark_code()
    else :
        pass

frame_combo = Frame(master , bd = 10 , bg = "#CA6B45")
frame_combo.place(relx = 0.5, rely = 0.5, anchor = CENTER)

combo = ttk.Combobox( frame_combo , values = var_list )
combo.set("Averages")
combo.pack(padx = 10 , pady = 10)

frame_button_select = Frame(master , bd = 12 , bg = "#CA6B45")
frame_button_select.place(x=560, y=100)

button_select = Button( frame_button_select, text = " Select File ", 
                        command = retrieve ,
                        fg = "black", font = "Times",
                        bg = "light blue" )
button_select.pack()

def empty_func() :
	# any code 
	pass
		
def open_func() :
	print("Open")
	
def save_func() :
	print("Save")
	
def exit_func() :
	print("Exit")
	master.destroy()

def find_func() :
	print("Find")
	
def debugger_func() :
	print("Debugger")
	
def replace_func() :
	print("Replace")

def documentation_func() :
	print("Documentations")
	
# main menu
main_menu = Menu(master)

# menu 1
file_menu = Menu( master , tearoff = 0 )
file_menu.add_command( label = "Open" , command = open_func )									
file_menu.add_command( label = "Save" , command = save_func )
file_menu.add_separator()
file_menu.add_command( label = "Exit" , command = exit_func )

main_menu.add_cascade( label = "File" , menu = file_menu )

# menu 2
tool_menu = Menu( master , tearoff = 0 )
tool_menu.add_command( label = "Find" , command = find_func)									
tool_menu.add_command( label = "Debugger" , command = debugger_func)
tool_menu.add_command( label = "Replace" , command = replace_func)

main_menu.add_cascade( label = "Tools" , menu = tool_menu )

# menu 3
help_menu = Menu( master , tearoff = 0 )
help_menu.add_command( label = "Documentations" , command = documentation_func)									
main_menu.add_cascade( label = "Help" , menu = help_menu )
master.config(menu = main_menu)

master.title("RFMT v3 - Transistor")
master.mainloop()
